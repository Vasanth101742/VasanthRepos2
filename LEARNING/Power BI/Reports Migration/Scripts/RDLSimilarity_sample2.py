import os
import re
import xml.etree.ElementTree as ET
from itertools import combinations
from difflib import SequenceMatcher
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.formatting.rule import ColorScaleRule


# =========================================================
# CONFIGURATION
# =========================================================

RDL_FOLDER = r"C:\Paginated Reports Dump as on 24-Aug-2026\Accounts"

OUTPUT_EXCEL = os.path.join(
    RDL_FOLDER,
    "RDL_Similarity_Report_Accounts.xlsx"
)

SIMILARITY_THRESHOLD = 80


# =========================================================
# READ AND NORMALIZE RDL
# =========================================================

def read_rdl(file_path):

    try:

        tree = ET.parse(file_path)
        root = tree.getroot()

        # Remove XML namespaces
        for element in root.iter():

            if "}" in element.tag:
                element.tag = element.tag.split("}", 1)[1]

        # Convert XML to string
        xml_text = ET.tostring(
            root,
            encoding="unicode"
        )

        # Normalize whitespace
        xml_text = re.sub(
            r"\s+",
            " ",
            xml_text
        )

        return xml_text.strip()

    except Exception as e:

        print(
            f"Error reading {file_path}: {e}"
        )

        return None


# =========================================================
# CALCULATE SIMILARITY
# =========================================================

def calculate_similarity(text1, text2):

    similarity = SequenceMatcher(
        None,
        text1,
        text2
    ).ratio()

    return similarity * 100


# =========================================================
# DETERMINE RESULT
# =========================================================

def get_result(similarity):

    if similarity >= 80:
        return "HIGH"

    elif similarity >= 60:
        return "MEDIUM"

    else:
        return "LOW"


# =========================================================
# CREATE EXCEL REPORT
# =========================================================

def create_excel_report(results):

    workbook = Workbook()

    worksheet = workbook.active

    worksheet.title = "RDL Similarity"


    # -----------------------------------------------------
    # Headers
    # -----------------------------------------------------

    headers = [
        "Report 1",
        "Report 2",
        "Similarity %",
        "Result"
    ]

    worksheet.append(headers)


    # -----------------------------------------------------
    # Header formatting
    # -----------------------------------------------------

    header_fill = PatternFill(
        start_color="1F4E78",
        end_color="1F4E78",
        fill_type="solid"
    )

    header_font = Font(
        color="FFFFFF",
        bold=True
    )

    for cell in worksheet[1]:

        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(
            horizontal="center"
        )


    # -----------------------------------------------------
    # Add data
    # -----------------------------------------------------

    for result in results:

        worksheet.append([
            result["Report 1"],
            result["Report 2"],
            result["Similarity %"],
            result["Result"]
        ])


    # -----------------------------------------------------
    # Format similarity column
    # -----------------------------------------------------

    for row in range(2, worksheet.max_row + 1):

        worksheet.cell(
            row=row,
            column=3
        ).number_format = "0.00"


    # -----------------------------------------------------
    # Conditional formatting
    # -----------------------------------------------------

    worksheet.conditional_formatting.add(
        f"C2:C{worksheet.max_row}",
        ColorScaleRule(
            start_type="num",
            start_value=0,
            start_color="F8696B",

            mid_type="num",
            mid_value=60,
            mid_color="FFEB84",

            end_type="num",
            end_value=100,
            end_color="63BE7B"
        )
    )


    # -----------------------------------------------------
    # Format HIGH/MEDIUM/LOW
    # -----------------------------------------------------

    high_fill = PatternFill(
        start_color="C6EFCE",
        end_color="C6EFCE",
        fill_type="solid"
    )

    medium_fill = PatternFill(
        start_color="FFEB9C",
        end_color="FFEB9C",
        fill_type="solid"
    )

    low_fill = PatternFill(
        start_color="FFC7CE",
        end_color="FFC7CE",
        fill_type="solid"
    )


    for row in range(2, worksheet.max_row + 1):

        result_cell = worksheet.cell(
            row=row,
            column=4
        )

        if result_cell.value == "HIGH":

            result_cell.fill = high_fill

        elif result_cell.value == "MEDIUM":

            result_cell.fill = medium_fill

        else:

            result_cell.fill = low_fill


    # -----------------------------------------------------
    # Column widths
    # -----------------------------------------------------

    worksheet.column_dimensions["A"].width = 40
    worksheet.column_dimensions["B"].width = 40
    worksheet.column_dimensions["C"].width = 18
    worksheet.column_dimensions["D"].width = 15


    # -----------------------------------------------------
    # Freeze header
    # -----------------------------------------------------

    worksheet.freeze_panes = "A2"


    # -----------------------------------------------------
    # Auto filter
    # -----------------------------------------------------

    worksheet.auto_filter.ref = worksheet.dimensions


    # -----------------------------------------------------
    # Save
    # -----------------------------------------------------

    workbook.save(OUTPUT_EXCEL)


# =========================================================
# MAIN
# =========================================================

def main():

    print("=" * 70)
    print("RDL REPORT SIMILARITY CHECKER")
    print("=" * 70)


    # -----------------------------------------------------
    # Find RDL files
    # -----------------------------------------------------

    rdl_files = [

        os.path.join(
            RDL_FOLDER,
            file
        )

        for file in os.listdir(RDL_FOLDER)

        if file.lower().endswith(".rdl")
    ]


    if len(rdl_files) < 2:

        print(
            "\nAt least 2 RDL files are required."
        )

        return


    print(
        f"\nFound {len(rdl_files)} RDL reports."
    )


    # -----------------------------------------------------
    # Read reports
    # -----------------------------------------------------

    reports = {}


    for file_path in rdl_files:

        print(
            "Reading:",
            os.path.basename(file_path)
        )

        content = read_rdl(file_path)

        if content:

            reports[file_path] = content


    # -----------------------------------------------------
    # Compare reports
    # -----------------------------------------------------

    results = []


    print("\nComparing reports...\n")


    for file1, file2 in combinations(
        reports.keys(),
        2
    ):

        similarity = calculate_similarity(
            reports[file1],
            reports[file2]
        )


        result = get_result(
            similarity
        )


        results.append({

            "Report 1":
                os.path.basename(file1),

            "Report 2":
                os.path.basename(file2),

            "Similarity %":
                round(similarity, 2),

            "Result":
                result
        })


    # -----------------------------------------------------
    # Sort highest similarity first
    # -----------------------------------------------------

    results.sort(
        key=lambda x: x["Similarity %"],
        reverse=True
    )


    # -----------------------------------------------------
    # Create Excel
    # -----------------------------------------------------

    create_excel_report(
        results
    )


    # -----------------------------------------------------
    # Display results
    # -----------------------------------------------------

    print("\n")
    print("=" * 90)

    print(
        f"{'REPORT 1':35}"
        f"{'REPORT 2':35}"
        f"{'SIMILARITY':15}"
        f"{'RESULT':10}"
    )

    print("=" * 90)


    for result in results:

        print(
            f"{result['Report 1'][:34]:35}"
            f"{result['Report 2'][:34]:35}"
            f"{result['Similarity %']:>10.2f}%     "
            f"{result['Result']}"
        )


    print("\n")
    print("=" * 70)

    print(
        "Excel report created:"
    )

    print(
        OUTPUT_EXCEL
    )

    print("=" * 70)


# =========================================================
# PROGRAM ENTRY
# =========================================================

if __name__ == "__main__":

    main()
